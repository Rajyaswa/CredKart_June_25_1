#====================Questions on Text File Handling===========================================
# 1.Write a program to create a new text file and write the string "Hello, World!" into it.
from _pyrepl import readline

print("\nQuestion1:Write a program to create a new text file and write the string \"Hello, World!\" into it.")
file=open(r"C:\Users\SMAK\PycharmProjects\PythonProject1\File_Handling\Text_File\File3.txt","w")
file.write("\"Hello, World!\"")
file.close()
# 2.Read the contents of a text file and print them line by line.
print("\nQuestion2:Read the contents of a text file and print them line by line.")
file=open(r"C:\Users\SMAK\PycharmProjects\PythonProject1\File_Handling\Text_File\File3.txt","r")
print(f"file contain:{file.read()}")
file.close()
# 3.Append the string "Python is fun!" to an existing text file.
print("\nQuestion3:Append the string \"Python is fun!\" to an existing text file.")
file=open(r"C:\Users\SMAK\PycharmProjects\PythonProject1\File_Handling\Text_File\File3.txt","a")
print(f"append file an existing text file:{file.write(" \n   \t \"Python is fun!\"")}")
file=open(r"C:\Users\SMAK\PycharmProjects\PythonProject1\File_Handling\Text_File\File3.txt","r")
print(f"read appended existing file:{file.read()}")
file.close()
# 4.Write a program to count the number of words in a text file.
print("\nQuestion4:Write a program to count the number of words in a text file.")
file=open(r"C:\Users\SMAK\PycharmProjects\PythonProject1\File_Handling\Text_File\File3.txt","r")
print(f"read appended existing file:{file.read()}")

# 5.Write a program to count the number of lines in a text file.
with open(r"C:\Users\SMAK\PycharmProjects\PythonProject1\File_Handling\Text_File\File3.txt","r")as d:
    lines = len(d.readlines())
    print('Total Number of lines:', lines)
file.close()
# 6.Create a text file and write the multiplication table of 5 (1 to 10) into it.
print("\nQuestion6:Create a text file and write the multiplication table of 5 (1 to 10) into it.")
for i in str(5): # 2 to 10
    file_name = "Table_Of_" + str(5) + ".txt"
    file = open(r"C:\Users\SMAK\PycharmProjects\PythonProject1\File_Handling\Text_File\File3.txt" + file_name,"w")
    for j in range (1,11):
        file.write(str(5) + " * " + str(j) + " = " + str(5 * j) + "\n")
    file.close()
# 7.Write a program to read a file and count the occurrences of a specific word.
file=open(r"C:\Users\SMAK\PycharmProjects\PythonProject1\File_Handling\Text_File\File3.txt","r")
print(f"read appended existing file:{file.read()}")

# 8.Open a text file in append mode and add the current date and time at the end.
file=open(r"C:\Users\SMAK\PycharmProjects\PythonProject1\File_Handling\Text_File\File3.txt","a")

# 9.Create a text file and write a list of names into it, each on a new line.
# 10.Write a program to reverse the contents of a text file and save it to a new file.
# 11.Copy the contents of one text file to another file.
# 12.Write a program to delete all the content in a text file.
# 13.Create a text file and write the first 10 Fibonacci numbers into it.
# 14.Write a program to find the longest line in a text file.
# 15.Create a program to merge the contents of two text files into a third file.
# 16.Write a program to replace a specific word in a text file with another word.
# 17.Write a program to check if a text file exists; if not, create it.
# 18.Create a text file and write all even numbers from 1 to 50 into it.
# 19.Write a program to count the number of uppercase letters in a text file.
# 20.Read a text file and print all lines that contain the word "Python.
# ==========================="Questions on Excel File Handling=======================
# 1.Create a new Excel file and write the header row "Name", "Age", and "City."
import openpyxl
file_path  = r"C:\Users\SMAK\PycharmProjects\PythonProject1\File_Handling\Excel_File\File1.xlsx"
sheet_name = "Sheet1"
Excel_File = openpyxl.load_workbook(file_path) # giving the file to our program #1
Sheet = Excel_File[sheet_name] # giving the sheet to our program #2
read_data1  = Sheet.cell(row=1, column=1).value #3
read_data2  = Sheet.cell(row=2, column=1).value #3

read_data = Sheet.cell(row=3, column=1).value

print(f" Data at row 1 and column 1 is : {read_data1}")
print(f" Data at row 2 and column 1 is : {read_data2}")
print(f" Data at row 1 and column 1 is : {read_data}")
print(f" Data at row 1 and column 2 is : {Sheet.cell(row=1, column=2).value}")
print(f"Total number of rows in sheet is : {Sheet.max_row}")

# 2.Write a program to add a row of data to an existing Excel sheet.
# 3.Read the data from the first row and column of an Excel file.
# 4.Write a program to count the number of rows in an Excel sheet.
# 5.Create an Excel sheet and write numbers 1 to 10 in the first column.
# 6.Read all the data from an Excel sheet and print it.
# 7.Write a program to update the value of a specific cell in an Excel file.
# 8.Create a new Excel sheet and write a multiplication table of 2 in the first column.
